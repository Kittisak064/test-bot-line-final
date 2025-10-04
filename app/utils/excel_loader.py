import os, re
from typing import Dict, List, Any, Tuple
from openpyxl import load_workbook

EXCEL_PATH = os.getenv("EXCEL_PATH", "./app/data/อีกครั้ง.xlsx")

# ---------- Sheet Names (ตามที่คุณใช้อยู่) ----------
SHEET_COMPANY      = "ข้อมูลบริษัท"
SHEET_PRODUCTS     = "ข้อมูลสินค้าและราคา"
SHEET_PERSONA      = "บุคลิกน้อง A.I."
SHEET_FAQ          = "FAQ"
SHEET_TRAIN_DOC    = "Training Doc"
SHEET_ORDERS       = "Orders"
SHEET_INTENT_PRE   = "Intent Instruction – ก่อนขาย"
SHEET_INTENT_POST  = "Intent Instruction – หลังการขาย"
SHEET_TRAINING_RAW = "Training Data"
SHEET_SYS_CONFIG   = "System Config"

# ---------- Helpers ----------
def _safe_cell(v):
    return "" if v is None else str(v).strip()

def _load_wb() -> Any:
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"ไม่พบไฟล์ Excel ที่ {EXCEL_PATH}")
    return load_workbook(EXCEL_PATH, data_only=True)

def _sheet(wb, name: str):
    if name not in wb.sheetnames:
        return None
    return wb[name]

# ---------- Loaders ----------
def load_company_info() -> Dict[str, str]:
    """คอลัมน์ A=หัวข้อ, B=รายละเอียด"""
    wb = _load_wb()
    ws = _sheet(wb, SHEET_COMPANY)
    if not ws: return {}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        key, val = (_safe_cell(r[0]), _safe_cell(r[1]) if len(r)>1 else "")
        if key: out[key] = val
    return out

def load_persona() -> Dict[str, str]:
    """คอลัมน์ A=หัวข้อ, B=รายละเอียด"""
    wb = _load_wb()
    ws = _sheet(wb, SHEET_PERSONA)
    if not ws: return {}
    out={}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        k, v = _safe_cell(r[0]), _safe_cell(r[1]) if len(r)>1 else ""
        if k: out[k]=v
    return out

def load_system_config() -> Dict[str, str]:
    """คอลัมน์ A=หัวข้อ, B=ค่า/การตั้งค่า"""
    wb = _load_wb()
    ws = _sheet(wb, SHEET_SYS_CONFIG)
    if not ws: return {}
    conf={}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        k, v = _safe_cell(r[0]), _safe_cell(r[1]) if len(r)>1 else ""
        if k: conf[k]=v
    return conf

def load_faq() -> List[Dict[str,str]]:
    """A=คำถาม, B=คำตอบ, C=หมวดหมู่, D=คีย์เวิร์ด"""
    wb=_load_wb()
    ws=_sheet(wb, SHEET_FAQ)
    if not ws: return []
    data=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        data.append({
            "question": _safe_cell(r[0]),
            "answer":   _safe_cell(r[1]),
            "category": _safe_cell(r[2]) if len(r)>2 else "",
            "keywords": _safe_cell(r[3]) if len(r)>3 else "",
        })
    return data

def load_training_doc() -> List[Dict[str,str]]:
    """A=หัวข้อ, B=รายละเอียด (C=รูป ถ้ามีไม่ใช้ตอนนี้)"""
    wb=_load_wb()
    ws=_sheet(wb, SHEET_TRAIN_DOC)
    if not ws: return []
    data=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        data.append({"title": _safe_cell(r[0]), "content": _safe_cell(r[1]) if len(r)>1 else ""})
    return data

def load_training_raw() -> str:
    """Training Data: รวมข้อความคอลัมน์ B ทั้งชีท"""
    wb=_load_wb()
    ws=_sheet(wb, SHEET_TRAINING_RAW)
    if not ws: return ""
    lines=[]
    for r in ws.iter_rows(min_row=1, values_only=True):
        if not r: continue
        # สมมติคอนเทนต์อยู่คอลัมน์ B
        lines.append(_safe_cell(r[1]) if len(r)>1 else "")
    return "\n".join([x for x in lines if x])

def load_intent_instructions() -> Dict[str, List[Dict[str,str]]]:
    """
    อ่านก่อนขาย/หลังการขาย
    Layout: A=บริบท/สถานการณ์, B=ให้ตอบแบบไหน, C=มี CTA ด้วยมั้ย
    """
    wb=_load_wb()
    out={}
    for sheet_name, key in [(SHEET_INTENT_PRE,"pre_sale"), (SHEET_INTENT_POST,"post_sale")]:
        ws=_sheet(wb, sheet_name)
        items=[]
        if ws:
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r: continue
                items.append({
                    "context": _safe_cell(r[0]),
                    "how_to":  _safe_cell(r[1]) if len(r)>1 else "",
                    "cta":     _safe_cell(r[2]) if len(r)>2 else "",
                })
        out[key]=items
    return out

def load_products() -> List[Dict[str, Any]]:
    """
    ชีท ‘ข้อมูลสินค้าและราคา’
    A=รหัสสินค้าในระบบขาย
    B=ชื่อสินค้าในระบบขาย
    C=ชื่อสินค้าที่มักถูกเรียก (คีย์เวิร์ด, คำพ้อง, เว้นวรรค/คอมมา)
    D=ขนาด
    E=หน่วย
    F=ราคาเต็ม
    G=ราคาขาย
    H=ราคาค่าขนส่ง
    I=หมวดหมู่
    """
    wb=_load_wb()
    ws=_sheet(wb, SHEET_PRODUCTS)
    if not ws: return []
    out=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        out.append({
            "sku":     _safe_cell(r[0]),
            "name":    _safe_cell(r[1]),
            "aliases": _safe_cell(r[2]),
            "size":    _safe_cell(r[3]),
            "unit":    _safe_cell(r[4]),
            "list_price": _safe_cell(r[5]),
            "sale_price": _safe_cell(r[6]),
            "shipping":   _safe_cell(r[7]),
            "category":   _safe_cell(r[8]) if len(r)>8 else "",
        })
    return out

# ---------- Simple Search ----------
def _tokenize(s: str) -> List[str]:
    return re.findall(r"[ก-๛A-Za-z0-9]+", s or "")

def match_products_by_query(products, q: str, top_k=3) -> List[Dict[str,Any]]:
    tokens = set(_tokenize(q.lower()))
    scored=[]
    for p in products:
        hay = " ".join([p["name"], p["aliases"]]).lower()
        score = sum(1 for t in tokens if t in hay)
        if score>0:
            scored.append((score, p))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [p for _,p in scored[:top_k]]

def match_faq_by_query(faqs, q: str, top_k=3) -> List[Dict[str,str]]:
    tokens = set(_tokenize(q.lower()))
    scored=[]
    for f in faqs:
        hay = " ".join([f["question"], f["keywords"]]).lower()
        score = sum(1 for t in tokens if t in hay)
        if score>0:
            scored.append((score, f))
    scored.sort(key=lambda x:x[0], reverse=True)
    return [f for _,f in scored[:top_k]]

# ---------- Aggregate ----------
def load_all() -> Dict[str, Any]:
    return {
        "company": load_company_info(),
        "persona": load_persona(),
        "sysconf": load_system_config(),
        "faq":     load_faq(),
        "training_doc": load_training_doc(),
        "training_raw": load_training_raw(),
        "intents": load_intent_instructions(),
        "products": load_products(),
    }

# พรีโหลดตอนสตาร์ท (รอบเดียว)
DATA = load_all()
